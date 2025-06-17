import os
import google.generativeai as genai
from flask import Flask, render_template, request, send_from_directory, jsonify, abort, after_this_request
from openpyxl import load_workbook
import uuid
import time
import math
import json
import sys
import threading
import traceback
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# --- 設定區 ---
# 建立 Flask 應用
app = Flask(__name__)

# 從環境變數讀取密碼，如果未設定，則使用預設值
AUTH_PASSWORD = os.getenv("AUTH_PASSWORD", "123")
UPLOAD_FOLDER = 'uploads'
DOWNLOAD_FOLDER = 'downloads'
GLOSSARY_FILE = 'dic.json'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER

# --- Email SMTP 設定 (從環境變數讀取) ---
SMTP_HOST = os.getenv("SMTP_HOST")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587))
SMTP_USER = os.getenv("SMTP_USER")      # SendGrid 登入使用者名稱 (應為 'apikey')
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")  # SendGrid API 金鑰 (應為 SG. 開頭的金鑰)
# --- 【最終修正】新增 SENDER_EMAIL 變數，與登入帳號分開 ---
SENDER_EMAIL = os.getenv("SENDER_EMAIL")    # 您在 SendGrid 上驗證過的寄件人信箱

# --- 非同步任務儲存區 ---
jobs = {}

# --- 載入專業詞彙表 ---
def load_glossary(filepath):
    """從 JSON 檔案載入專業詞彙表。"""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        # 將越南文轉為小寫以進行不區分大小寫的比對
        glossary = {item['vietnamese'].lower(): item['chinese'] for item in data}
        print(f"✅ 成功載入專業詞彙表，共 {len(glossary)} 條目。")
        return glossary
    except Exception as e:
        print(f"❌ 載入詞彙表 '{filepath}' 時發生錯誤: {e}")
        return None

VIET_TO_CHI_GLOSSARY = load_glossary(GLOSSARY_FILE)

# --- 設定 Gemini API 金鑰 ---
try:
    gemini_api_key = os.getenv("GEMINI_API_KEY")
    if not gemini_api_key:
        raise ValueError("錯誤：請設定 GEMINI_API_KEY 環境變數")
    genai.configure(api_key=gemini_api_key)
    model = genai.GenerativeModel('gemini-1.5-flash-latest')
except Exception as e:
    print(f"初始化 Gemini 模型時發生錯誤: {e}")
    model = None

# 確保暫存資料夾存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)


# --- 寄送電子郵件的函式 (已修正) ---
def send_email_with_attachment(recipient_email, subject, body, file_path):
    """使用設定好的 SMTP 資訊寄送帶有附件的郵件。"""
    # 檢查所有郵件相關設定是否完整
    if not all([SMTP_HOST, SMTP_USER, SMTP_PASSWORD, SENDER_EMAIL]):
        print("警告：SMTP 電子郵件環境變數設定不完整，已跳過郵件寄送步驟。")
        return False
    
    try:
        msg = MIMEMultipart()
        # --- 【最終修正】寄件人地址使用新的 SENDER_EMAIL 變數 ---
        msg['From'] = SENDER_EMAIL 
        msg['To'] = recipient_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        with open(file_path, "rb") as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(file_path))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        msg.attach(part)

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            # 【最終修正】登入時仍然使用 'apikey' 和 API Key
            server.login(SMTP_USER, SMTP_PASSWORD) 
            server.send_message(msg)
        
        print(f"✅ 成功將翻譯檔案寄送至: {recipient_email}")
        return True
    except Exception as e:
        print(f"❌ 寄送郵件至 {recipient_email} 時發生錯誤: {e}")
        traceback.print_exc()
        return False


# --- 核心翻譯功能 (保持不變) ---
def translate_text_batch_with_gemini(texts, separator):
    if not texts or not model: return texts
    combined_text = separator.join(texts)
    relevant_terms_list = []
    if VIET_TO_CHI_GLOSSARY:
        source_text_lower = combined_text.lower()
        for viet_term, chi_translation in VIET_TO_CHI_GLOSSARY.items():
            if viet_term in source_text_lower:
                relevant_terms_list.append(f"- 越南原文 '{viet_term}' 應翻譯為 '{chi_translation}'")
    glossary_section = "**【優先翻譯詞彙】**\n無特定術語需要優先處理。"
    if relevant_terms_list:
        glossary_section = ("**【優先翻譯詞彙】**\n" "在翻譯時，請務必遵循以下術語對照，優先使用指定的翻譯：\n" + "\n".join(relevant_terms_list))
    prompt = f"""
    任務：請將以下由特殊分隔符 `"{separator}"` 串連起來的越南文文字，逐一精準地翻譯成繁體中文。
    規則：
    1.  翻譯完成後，必須使用完全相同的分隔符 `"{separator}"` 將所有翻譯結果重新串連起來。
    2.  除了翻譯後的文字和分隔符外，不要包含任何前言、解釋或額外字元。
    3.  如果某個片段是數字、英文、純粹的標點符號或看起來不是越南文，請直接返回該片段原文。
    4.  保持片段的順序與原文完全一致。
    5.  返回的片段數量必須與原始片段數量完全相同。
    {glossary_section}
    ---
    待翻譯的原文組合:
    "{combined_text}"
    ---
    翻譯後的繁體中文組合:
    """
    for i in range(3):
        try:
            response = model.generate_content(prompt)
            if response.text:
                translated_texts = response.text.strip().split(separator)
                if len(translated_texts) == len(texts): return translated_texts
                else: print(f"批次錯誤：API回傳的片段數量 ({len(translated_texts)}) 與原始數量 ({len(texts)}) 不符。"); return texts
            else: print(f"警告: Gemini 為批次任務回傳了空的結果。"); return texts
        except Exception as e:
            if '429' in str(e): wait_time = 2**i; print(f"觸發 API 速率限制，等待 {wait_time} 秒後重試..."); time.sleep(wait_time)
            else: print(f"呼叫 Gemini API 時發生錯誤: {e}"); return texts
    print("警告: 重試多次後，批次翻譯仍然失敗。將使用原文填充。"); return texts

def process_excel_file_optimized(input_path, output_path):
    workbook = load_workbook(input_path)
    SEPARATOR = "|||$$$|||"
    CHUNK_SIZE = 150
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        texts_to_translate, cell_locations = [], []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str) and cell.value.strip():
                    texts_to_translate.append(cell.value)
                    cell_locations.append((row_idx, col_idx))
        if not texts_to_translate: continue
        all_translated_texts = []
        total_chunks = math.ceil(len(texts_to_translate) / CHUNK_SIZE)
        for i in range(0, len(texts_to_translate), CHUNK_SIZE):
            chunk = texts_to_translate[i:i + CHUNK_SIZE]
            current_chunk_num = (i // CHUNK_SIZE) + 1
            print(f"工作表 '{sheet_name}': 正在處理第 {current_chunk_num} / {total_chunks} 批...")
            translated_chunk = translate_text_batch_with_gemini(chunk, SEPARATOR)
            if len(translated_chunk) != len(chunk): print(f"警告：第 {current_chunk_num} 批翻譯失敗，此批次將使用原文。"); all_translated_texts.extend(chunk)
            else: all_translated_texts.extend(translated_chunk)
        if len(all_translated_texts) == len(cell_locations):
            for i, location in enumerate(cell_locations):
                row, col = location
                original_text, translated_text = texts_to_translate[i], all_translated_texts[i].strip()
                if original_text.strip() != translated_text and translated_text: sheet.cell(row=row, column=col).value = f"{original_text}\n{translated_text}"
                else: sheet.cell(row=row, column=col).value = original_text
        else: print(f"【嚴重錯誤】：最終文本數量與位置數量不符，工作表 '{sheet_name}' 不進行任何修改。")
    workbook.save(output_path)

# --- 背景處理函式 ---
def process_file_in_background(job_id, input_path, output_path):
    global jobs
    try:
        print(f"背景任務 {job_id} 開始處理檔案: {os.path.basename(input_path)}")
        process_excel_file_optimized(input_path, output_path)
        jobs[job_id]['status'] = 'completed'
        jobs[job_id]['download_url'] = f'/download/{os.path.basename(output_path)}'
        print(f"背景任務 {job_id} 成功完成。")
        job_info = jobs.get(job_id)
        if job_info and job_info.get('email'):
            recipient = job_info['email']
            filename = os.path.basename(output_path)
            subject = f"您的 Excel 翻譯任務已完成 | {filename}"
            body = f"您好，\n\n您上傳的檔案 ({filename}) 已成功翻譯完成。\n請查收附件。\n\n此為系統自動發送的郵件，請勿直接回覆。"
            send_email_with_attachment(recipient, subject, body, output_path)
    except Exception as e:
        print(f"背景任務 {job_id} 發生嚴重錯誤: {e}")
        traceback.print_exc()
        jobs[job_id]['status'] = 'failed'
        jobs[job_id]['error'] = str(e)
    finally:
        if os.path.exists(input_path):
            os.remove(input_path)
            print(f"已刪除任務 {job_id} 的原始上傳檔案: {os.path.basename(input_path)}")

# --- Web 路由 ---
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file_async():
    password = request.form.get('password')
    if password != AUTH_PASSWORD:
        return jsonify({'success': False, 'error': '密碼錯誤！'}), 401
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '沒有檔案被上傳。'}), 400
    file = request.files['file']
    if file.filename == '' or not file.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'error': '請選擇一個 .xlsx 格式的檔案。'}), 400
    if not model:
        return jsonify({'success': False, 'error': '伺服器端 Gemini 模型未成功初始化，請聯絡管理員。'}), 500
    user_email = request.form.get('email', None)
    unique_id = str(uuid.uuid4())
    original_filename = f"{unique_id}_original.xlsx"
    translated_filename = f"{unique_id}_translated.xlsx"
    input_path = os.path.join(app.config['UPLOAD_FOLDER'], original_filename)
    output_path = os.path.join(app.config['DOWNLOAD_FOLDER'], translated_filename)
    file.save(input_path)
    job_id = unique_id
    jobs[job_id] = {'status': 'processing', 'email': user_email}
    thread = threading.Thread(target=process_file_in_background, args=(job_id, input_path, output_path))
    thread.daemon = True
    thread.start()
    print(f"已成功接收檔案並創建背景任務 {job_id}。使用者郵箱: {user_email or '未提供'}")
    return jsonify({'success': True, 'job_id': job_id})

@app.route('/status/<job_id>')
def get_status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({'status': 'not_found', 'error': '找不到指定的任務ID，可能伺服器已重啟。'}), 404
    return jsonify(job)

@app.route('/download/<filename>')
def download_file(filename):
    file_path = os.path.join(app.config['DOWNLOAD_FOLDER'], filename)
    if not os.path.isfile(file_path):
        abort(404, "找不到檔案，可能已被刪除或處理失敗。")
    @after_this_request
    def cleanup(response):
        job_id = filename.replace("_translated.xlsx", "")
        if job_id in jobs:
            del jobs[job_id]
            print(f"已從任務列表中移除已完成的任務: {job_id}")
        try:
            os.remove(file_path)
            print(f"已刪除已下載的檔案: {filename}")
        except OSError as e:
            print(f"刪除檔案時出錯: {e.strerror}")
        return response
    return send_from_directory(app.config['DOWNLOAD_FOLDER'], filename, as_attachment=True)

# --- 啟動伺服器 ---
if __name__ == '__main__':
    if not model or not VIET_TO_CHI_GLOSSARY:
        print("致命錯誤：Gemini 模型或專業詞彙表未能成功載入。請檢查設定後再重新啟動。")
        sys.exit(1)
    # 啟動時檢查郵件設定，並給予提示
    if not all([SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD, SENDER_EMAIL]):
        print("警告：部分或全部郵件服務環境變數未設定。郵件功能可能無法使用。")
    print("Gemini 模型及專業詞彙表已成功載入。伺服器準備就緒。")
    app.run(host='0.0.0.0', port=os.environ.get('PORT', 5000), debug=False)
